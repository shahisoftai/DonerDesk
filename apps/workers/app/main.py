"""DonorDesk async workers (FastAPI).

Phase 1 architecture:
  - The NestJS/Next API handles most synchronous flows.
  - Long-running or streaming AI tasks (OCR, embedding, chunked ingestion) are
    meant to be invoked through Kestra flows.
  - This service exposes an HTTP API that Kestra flows call. For Phase 1 it
    simply re-exports the same deterministic stub strategies used in the
    TypeScript API, so swapping Kestra in later doesn't change the contract.

Routes:
  GET  /health
  POST /v1/parse          — parse a document (multipart upload)
  POST /v1/suggest-tags   — suggest evidence tags
  POST /v1/polish         — polish activity narrative
  POST /v1/draft-section  — draft a report section (stub)
"""
from __future__ import annotations

import io
from typing import Any

from fastapi import FastAPI, File, HTTPException, UploadFile
from pydantic import BaseModel, Field

app = FastAPI(title="DonorDesk Workers", version="0.1.0")


SENSITIVE_KEYWORDS = [
    "beneficiary name",
    "phone",
    "cnic",
    "national id",
    "passport",
    "children",
    "medical",
    "diagnosis",
    "patient",
]


@app.get("/health")
def health() -> dict[str, Any]:
    return {"status": "ok", "service": "donordesk-workers"}


@app.post("/v1/parse")
async def parse_document(file: UploadFile = File(...)) -> dict[str, Any]:
    """Extract text from an uploaded file. Lightweight parsers only."""
    data = await file.read()
    filename = (file.filename or "upload.bin").lower()
    text = ""
    try:
        if filename.endswith(".txt") or (file.content_type or "").startswith("text/"):
            text = data.decode("utf-8", errors="ignore")
        elif filename.endswith(".csv"):
            text = data.decode("utf-8", errors="ignore")
        elif filename.endswith(".docx"):
            try:
                from docx import Document
                doc = Document(io.BytesIO(data))
                text = "\n".join(p.text for p in doc.paragraphs)
            except Exception as exc:  # pragma: no cover
                raise HTTPException(status_code=400, detail=f"docx parse failed: {exc}") from exc
        elif filename.endswith(".xlsx"):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
                parts = []
                for sheet in wb.worksheets:
                    parts.append(f"# {sheet.title}")
                    for row in sheet.iter_rows(values_only=True):
                        parts.append("\t".join(str(c) if c is not None else "" for c in row))
                text = "\n".join(parts)
            except Exception as exc:  # pragma: no cover
                raise HTTPException(status_code=400, detail=f"xlsx parse failed: {exc}") from exc
        elif filename.endswith(".pdf") or file.content_type == "application/pdf":
            try:
                from PyPDF2 import PdfReader
                reader = PdfReader(io.BytesIO(data))
                text = "\n".join(page.extract_text() or "" for page in reader.pages)
            except Exception:  # pragma: no cover
                text = ""
        else:
            text = filename
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"parse failed: {exc}") from exc
    return {"text": text, "filename": file.filename, "size": len(data)}


class SuggestTagsRequest(BaseModel):
    file_name: str
    file_type: str
    extracted_text: str | None = None
    existing_activities: list[dict[str, Any]] = Field(default_factory=list)
    existing_indicators: list[dict[str, Any]] = Field(default_factory=list)


@app.post("/v1/suggest-tags")
def suggest_tags(req: SuggestTagsRequest) -> dict[str, Any]:
    """Heuristic stub mirroring the TypeScript tagger so behaviour matches."""
    haystack = f"{req.file_name}\n{req.extracted_text or ''}".lower()
    keywords = {
        "attendance": ("ATTENDANCE_SHEET", "HIGH"),
        "photo": ("PHOTO", "HIGH"),
        "picture": ("PHOTO", "HIGH"),
        "distribution": ("DISTRIBUTION_LIST", "HIGH"),
        "training": ("TRAINING_RECORD", "MEDIUM"),
        "visit": ("FIELD_VISIT_REPORT", "MEDIUM"),
        "monitoring": ("MONITORING_REPORT", "MEDIUM"),
        "kobo": ("KOBO_ODK_EXPORT", "HIGH"),
        "odk": ("KOBO_ODK_EXPORT", "HIGH"),
        "procurement": ("PROCUREMENT_DOCUMENT", "MEDIUM"),
        "approval": ("APPROVAL_DOCUMENT", "MEDIUM"),
        "beneficiary": ("BENEFICIARY_LIST", "HIGH"),
        "minutes": ("MEETING_MINUTES", "MEDIUM"),
        "case": ("CASE_STUDY", "LOW"),
        "financial": ("FINANCIAL_DOCUMENT", "MEDIUM"),
        "invoice": ("FINANCIAL_DOCUMENT", "HIGH"),
        "supplier": ("SUPPLIER_DOCUMENT", "MEDIUM"),
    }
    tags: list[dict[str, Any]] = []
    matched = None
    for kw, (etype, conf) in keywords.items():
        if kw in haystack:
            if matched is None or conf == "HIGH":
                matched = (etype, conf)
    tags.append({
        "field": "evidenceType",
        "value": matched[0] if matched else "OTHER",
        "confidence": matched[1] if matched else "LOW",
        "accepted": False,
    })
    for activity in req.existing_activities[:5]:
        if activity.get("title", "").lower()[:6] in haystack:
            tags.append({"field": "activityId", "value": activity["id"], "confidence": "MEDIUM", "accepted": False})
            break
    for ind in req.existing_indicators[:5]:
        code = (ind.get("code") or "").lower()
        name = (ind.get("name") or "").lower()[:8]
        if code and code in haystack:
            tags.append({"field": "indicatorId", "value": ind["id"], "confidence": "MEDIUM", "accepted": False})
            break
        if name and name in haystack:
            tags.append({"field": "indicatorId", "value": ind["id"], "confidence": "MEDIUM", "accepted": False})
            break
    hits = [kw for kw in SENSITIVE_KEYWORDS if kw in haystack]
    warning = (
        "This file may contain sensitive personal data ("
        + ", ".join(hits[:3])
        + "). Please verify access level before sharing or exporting."
    ) if hits else None
    return {
        "summary": "Suggested classification based on filename and extracted text.",
        "tags": tags,
        "sensitivity_warning": warning,
        "model": "stub-v1",
    }


class PolishRequest(BaseModel):
    rough_summary: str
    achievements: str = ""
    challenges: str = ""
    lessons_learned: str = ""


@app.post("/v1/polish")
def polish(req: PolishRequest) -> dict[str, Any]:
    sentences = [
        "During the reporting period, the team implemented the planned activities.",
        f"Field notes: {req.rough_summary[:300]}",
    ]
    if req.achievements:
        sentences.append(f"Key achievements include: {req.achievements[:200]}.")
    if req.challenges:
        sentences.append(f"Challenges encountered: {req.challenges[:200]}.")
    if req.lessons_learned:
        sentences.append(f"Lessons learned: {req.lessons_learned[:200]}.")
    return {"narrative": " ".join(sentences), "model": "stub-v1"}


class DraftSectionRequest(BaseModel):
    section_title: str
    project_name: str
    donor_name: str
    report_type: str
    template_sections: list[dict[str, Any]] = Field(default_factory=list)
    logframe_summary: str = ""
    indicator_summary: list[dict[str, Any]] = Field(default_factory=list)
    activities: list[dict[str, Any]] = Field(default_factory=list)
    evidence_by_activity: dict[str, list[dict[str, Any]]] = Field(default_factory=dict)
    evidence_by_indicator: dict[str, list[dict[str, Any]]] = Field(default_factory=dict)


@app.post("/v1/draft-section")
def draft_section(req: DraftSectionRequest) -> dict[str, Any]:
    title = req.section_title.lower()
    if "executive summary" in title:
        content = (
            f"This {req.report_type.lower().replace('_', ' ')} report summarises the implementation "
            f"progress of {req.project_name} funded by {req.donor_name}. "
            f"During the reporting period, {len(req.activities)} activities were completed."
        )
        refs = []
        if req.activities:
            first = req.activities[0]
            refs.append({"type": "activity", "id": first.get("id", ""), "label": first.get("title", "")})
        return {"title": req.section_title, "content": content, "source_references": refs, "unsupported_claims": []}
    if "indicator progress" in title:
        rows = [
            "| Code | Indicator | Baseline | Target | Period | Cumulative |",
            "| --- | --- | --- | --- | --- | --- |",
        ]
        for ind in req.indicator_summary:
            rows.append(
                f"| {ind.get('code', '')} | {ind.get('name', '')} | {ind.get('baseline', '')} | "
                f"{ind.get('target', '')} | {ind.get('periodAchievement', '')} {ind.get('unit', '') or ''} | "
                f"{ind.get('cumulativeAchievement', '')} {ind.get('unit', '') or ''} |"
            )
        content = "\n".join(rows) if len(rows) > 1 else "No indicators yet."
        refs = [
            {"type": "indicator", "id": ind.get("code", ""), "label": ind.get("name", "")}
            for ind in req.indicator_summary
        ]
        return {"title": req.section_title, "content": content, "source_references": refs, "unsupported_claims": []}
    if "annex" in title:
        refs = []
        for items in req.evidence_by_activity.values():
            for e in items:
                refs.append({"type": "evidence", "id": e.get("id", ""), "label": e.get("title", "")})
        return {
            "title": req.section_title,
            "content": "Refer to the evidence pack attached to this report for the full annex list.",
            "source_references": refs,
            "unsupported_claims": ["No evidence attached yet"] if not refs else [],
        }
    return {
        "title": req.section_title,
        "content": "[Section content will be drafted from logframe, indicators, activities, and evidence.]",
        "source_references": [],
        "unsupported_claims": [],
    }
