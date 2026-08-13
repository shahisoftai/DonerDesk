"""Document parsing for the DonorDesk workers.

Routing (which parser handles a file) is data-driven from the shared strategy
(`_strategy_data.py`), so it stays in lockstep with the TypeScript parser. The
per-format extraction logic lives here as small, single-responsibility helpers.
"""
from __future__ import annotations

import io
from typing import Any

from fastapi import HTTPException

from ._strategy_data import PARSE_ROUTES


def route_file(filename: str, content_type: str | None) -> str:
    """Return the parser handler name for a file (first route match wins)."""
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    ct = content_type or ""
    for route in PARSE_ROUTES:
        if ext in route["extensions"]:
            return str(route["handler"])
        if any(needle in ct for needle in route["contentTypeContains"]):
            return str(route["handler"])
    return "filename"


def _parse_text(data: bytes) -> str:
    return data.decode("utf-8", errors="ignore")


def _parse_docx(data: bytes) -> str:
    from docx import Document

    doc = Document(io.BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs)


def _parse_xlsx(data: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    parts: list[str] = []
    for sheet in wb.worksheets:
        parts.append(f"# {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            parts.append("\t".join(str(c) if c is not None else "" for c in row))
    return "\n".join(parts)


def _parse_pdf(data: bytes) -> str:
    from PyPDF2 import PdfReader

    reader = PdfReader(io.BytesIO(data))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def parse(data: bytes, filename: str, content_type: str | None) -> dict[str, Any]:
    """Parse a file's bytes into extracted text (used by `POST /v1/parse`).

    This is the module the Kestra flow references (`from app.parsers import parse`).
    """
    handler = route_file(filename, content_type)
    try:
        if handler == "filename":
            text = filename
        elif handler == "pdf":
            try:
                text = _parse_pdf(data)
            except Exception:  # pdf-parse is best-effort: fall back to empty text.
                text = ""
        elif handler == "docx":
            try:
                text = _parse_docx(data)
            except Exception as exc:
                raise HTTPException(status_code=400, detail=f"docx parse failed: {exc}") from exc
        elif handler == "xlsx":
            try:
                text = _parse_xlsx(data)
            except Exception as exc:
                raise HTTPException(status_code=400, detail=f"xlsx parse failed: {exc}") from exc
        else:
            text = _parse_text(data)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"parse failed: {exc}") from exc
    return {"text": text, "filename": filename, "size": len(data)}
